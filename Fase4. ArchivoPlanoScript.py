import pandas as pd
from time import time
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
from openpyxl.utils.dataframe import dataframe_to_rows
pd.options.mode.chained_assignment = None 
warnings.simplefilter(action='ignore', category=UserWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=RuntimeWarning)
from shareplum import Office365
from shareplum.site import Version
from shareplum import Site
import numpy as np
from datetime import date
import math
import tkinter as tk
from tkcalendar import Calendar

today = date.today()
start_time = time()
row_limit = 5000

authcookie = Office365('https://sunshinebouquet1.sharepoint.com/', username='scastro@sunshinebouquet.com', password='CCyl3uwWUK6ZD6sf').GetCookies()
siteAprovisionamiento = Site('https://sunshinebouquet1.sharepoint.com/sites/aprovisionamiento',version=Version.v2019, authcookie=authcookie)
siteDBLogistics = Site('https://sunshinebouquet1.sharepoint.com/sites/CosteodeTransporte',version=Version.v2019, authcookie=authcookie)

def get_excel_sh(site, folder1:str,folder2:str,folder3:str, namefile:str, sheetname:str,typeFolder:int):
#'Función para leer Excel Online Sharepoint'
    if typeFolder ==1:
        folder = site.Folder(f'Shared%20Documents/{folder1}/{folder2}')
    elif typeFolder==2:
        folder = site.Folder(f'Shared%20Documents/Indicadores/Agroquímicos/Resultados traslados/{folder1}/{folder2}')
    elif typeFolder==3:
        folder = site.Folder(f'Shared%20Documents/Indicadores/Agroquímicos/Resultados traslados/{folder1}/{folder2}/{folder3}')
    else:
        folder = site.Folder(f'Documentos%20compartidos/{folder1}')
    df = pd.read_excel(folder.get_file(namefile), sheet_name=sheetname)
    return df

def create_excel(df,namewb:str,namesh:str): #'Función para crear Excel Local'
    excel_st = Workbook()
    ws = excel_st.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    filename = os.path.dirname(__file__)+f'\\{namewb}.xlsx' 
    ws.title = namesh
    excel = excel_st.save(filename=filename)
    print('------------------ Excel Creado ------------------\n')
    return excel

def create_sheet(df,namewb:str,namesh:str):
    filename = os.path.dirname(__file__)+f'\\{namewb}.xlsx'
    writer = pd.ExcelWriter(filename, engine = 'openpyxl', mode='a', if_sheet_exists ='replace')
    df.to_excel(writer, sheet_name =namesh, index=False)
    writer._save()
    writer.close

def file_upload_to_sharepoint(siteExport,folderAño:str,folderSemana:str,fileName:str):
    fileNamePath= os.path.dirname(__file__)+f'\\{fileName}.xlsx'
    folder = siteExport.Folder(f'Shared%20Documents/Indicadores/Agroquímicos/Resultados traslados/{folderAño}/{folderSemana}')
    with open(fileNamePath, mode='rb') as file:
        fileContent = file.read()
    folder.upload_file(fileContent, f'{fileName}.xlsx')
    
def eliminarEspacios (x):
    x = x.strip()
    return x

# GUI of calendar
def createCalendarObject (title):
    ano = int(today.strftime("%Y"))
    mes = int(today.strftime("%m"))
    dia = int(today.strftime("%d"))
    tkobj = tk.Tk()
    tkobj.geometry("400x400")
    tkobj.title(title)
    tkc = Calendar(tkobj,selectmode = "day",year=ano,month=mes,date=dia)
    tkc.pack(pady=40)
    if title == "Fecha inicial":
        continueButton = tk.Button(tkobj, text="Seleccionar fecha", command=tkobj.destroy).pack()
    else: 
        continueButton = tk.Button(tkobj, text="Seleccionar", command=tkobj.destroy).pack()
    tkobj.mainloop()
    input = tkc.get_date()
    return input

#To change date to number format
def dateToString(date):
    año = slice(0,4)
    mes = slice(5,7)
    dia = slice(8,10)
    fecha = str(date[año]) + str(date[mes]) + str(date[dia])
    return fecha

aviso = print('Para generar el archivo de manera correcta se debe llenar la hoja de "Parámetros archivo plano" del Excel en línea llamado "Diccionarios".',end='\n'
              'Presione Enter si está seguro que estos datos ya están correctos en el sitio de Sharepoint.')
aviso2 = input("")

fechaDescarga = dateToString(str(today))
aviso = print('Ingrese: ',end='\n'
              '1: Si se va a correr el programa normal')
print(end='\n')
aviso3 = print('2: Si se van a realizar adicionales')      
adicionales = int(input("Escriba alguna de las 2 opciones anteriores: "))

parametros = get_excel_sh(siteAprovisionamiento,'Indicadores','Agroquímicos',"Parámetro",'Diccionarios.xlsx','Parámetros archivo plano',1)
año = parametros['Valor'][0]
semana = parametros['Valor'][1]
tipoDeDocumento = parametros['Valor'][2]
concepto = parametros['Valor'][3]
grupoDeClaseDeDocumento = parametros['Valor'][4]
claseDeDocumento = parametros['Valor'][5]
estadoDelDocumento = parametros['Valor'][6]
estadoDeImpresion = parametros['Valor'][7]
terceroComprador = parametros['Valor'][8]
sucursalDelProveedor = parametros['Valor'][9]
indicadorDeTasa = parametros['Valor'][10]
tasaDeConversion = parametros['Valor'][11]
monedaLocal = parametros['Valor'][12]
tasaLocal = parametros['Valor'][13]
indicadorDeContacto = parametros['Valor'][14]
motivo = parametros['Valor'][15]
indicadorDeObsequio = parametros['Valor'][16]
notas = parametros['Valor'][17]
detalle = parametros['Valor'][18]
descripcionDelItem = parametros['Valor'][19]
unidadNegocioMovimiento = parametros['Valor'][20]
tasaDescuentoCondicionado = parametros['Valor'][21]
descuentoGlobal1 = parametros['Valor'][22]
descuentoGlobal2 = parametros['Valor'][23]
monedaDelDocumento = parametros['Valor'][24]
monedaBaseDeConversion = parametros['Valor'][25]
solicitante = parametros['Valor'][26]
consecutivoDocumentoReferencia = int(parametros['Valor'][27])+1
fechaEntregaLejanias = parametros['Valor'][28]
fechaEntrega = parametros['Valor'][29]


#----------Archivo plano (cambiar por leer cada archivo del SP y concatenar dataframes)
folder = siteAprovisionamiento.Folder(f'Shared%20Documents/Indicadores/Agroquímicos/Resultados traslados/{año}/Semana{semana}')
foldersComprador = folder.folders

# Lista para almacenar los DataFrames de los archivos
dataframes = []
for folderI in foldersComprador:
    if adicionales==1:
        folderIndividual = siteAprovisionamiento.Folder(f'Shared%20Documents/Indicadores/Agroquímicos/Resultados traslados/{año}/Semana{semana}/{folderI}')
    else:
        folderIndividual = siteAprovisionamiento.Folder(f'Shared%20Documents/Indicadores/Agroquímicos/Resultados traslados/{año}/Semana{semana}/Adicionales/{folderI}')
    files = folderIndividual.files
    for file in files:
        if file['Name'].upper().startswith('IN'):
            if file['Name'].lower().endswith('.xlsx'):
                nombre = file['Name'].split('.')[0]
                if nombre == "Inventario disponible-faltante":
                    continue
                print(nombre)
                file_path = folderIndividual.get_file(file['Name'])
                df = pd.read_excel(file_path, sheet_name='Hoja1')
                dataframes.append(df)
                archivoPlano = pd.concat(dataframes)
            
archivoPlano = archivoPlano.reset_index()
archivoPlano['Concatenado'] = archivoPlano["Bodega"] + archivoPlano['Razón social proveedor']
archivoPlano = archivoPlano.sort_values(by = ['Bodega','Razón social proveedor','SisFinCode'], ascending = [True,True,True],ignore_index=True )
archivoPlanoCopia = archivoPlano.copy()

#------Ajustar los biológicos (los odio fucking bichos)
listBichitos = [3868,4709,6602,7484,7485,7731,8056,8057,8653,10941]
if semana%2==1:
    if adicionales==1:
        inventarioFaltante = get_excel_sh(siteAprovisionamiento,año,f'Semana{semana}','','Inventario disponible-faltante.xlsx','Hoja1',2)
    else:
        inventarioFaltante = get_excel_sh(siteAprovisionamiento,año,f'Semana{semana}','Adicionales','Inventario disponible-faltante.xlsx','Hoja1',3)
    
    inventarioFaltante = inventarioFaltante[inventarioFaltante['SisFinCode'].isin(listBichitos)]
    inventarioFaltante = inventarioFaltante[inventarioFaltante["Semanas de abastecimiento"] == 2]
    inventarioFaltante.rename(columns = {f'Consumo Semana ({semana+2})':f'{semana+2}', f'Consumo Semana ({semana+3})':f'{semana+3}'}, inplace = True)
    inventarioFaltante = inventarioFaltante.melt(id_vars=['Bodega','SisFinCode'], value_vars=[f'{semana+2}', f'{semana+3}'], var_name='Semana', value_name='Consumo')
    inventarioFaltante[['Semana']] = inventarioFaltante[["Semana"]].astype(int)
    archivoPlano = pd.merge(archivoPlano,inventarioFaltante,how='left',left_on= ['Bodega','SisFinCode'], right_on= ['Bodega','SisFinCode'])
    archivoPlano['Consumo'] = archivoPlano['Consumo'].fillna(0)
    archivoPlano['Unidades de compra'] = np.where(archivoPlano['Consumo'] == 0,archivoPlano['Unidades de compra'],archivoPlano['Consumo'])

####Hacer lógica de microorgnismos a dos semanas
indice = 0
consecutivo1 = 1
bodegaProveedorAnterior = archivoPlano['Concatenado'][0]
archivoPlano['consecutivoOC'] = 0
archivoPlano['consecutivoItem'] = 0
archivoPlano['documentoReferencia'] = consecutivoDocumentoReferencia
while indice<len(archivoPlano):
    bodegaProveedorActual = archivoPlano['Concatenado'][indice]
    if bodegaProveedorAnterior != bodegaProveedorActual:
        consecutivo1+=1
        consecutivoDocumentoReferencia+=1
    archivoPlano['consecutivoOC'][indice] = consecutivo1
    archivoPlano['documentoReferencia'][indice] = consecutivoDocumentoReferencia
    bodegaProveedorAnterior = bodegaProveedorActual
    indice+=1

indice = 0
consecutivo2 = 1
consecutivo1Anterior = archivoPlano['consecutivoOC'][0]
while indice<len(archivoPlano):
    consecutivo1Actual = archivoPlano['consecutivoOC'][indice]
    if consecutivo1Anterior != consecutivo1Actual:
        consecutivo2 = 1
    archivoPlano['consecutivoItem'][indice] = consecutivo2
    consecutivo2+=1
    consecutivo1Anterior = consecutivo1Actual
    indice+=1

archivoPlano = archivoPlano[['consecutivoOC','documentoReferencia','consecutivoItem','Bodega','SisFinCode','Uni','Razón social proveedor','Precio Actual Compra','UM Compras','Unidades de compra','Semana']]
archivoPlano['tipoDocumento'] = tipoDeDocumento
archivoPlano['Concepto'] = concepto
archivoPlano['Grupo de clase de documento'] = grupoDeClaseDeDocumento
archivoPlano['Clase de documento'] = claseDeDocumento
archivoPlano['Estado del documento'] = estadoDelDocumento
archivoPlano['Estado de impresión'] = estadoDeImpresion
archivoPlano['comprador'] = terceroComprador
archivoPlano['Sucursal del proveedor'] = sucursalDelProveedor
archivoPlano['Indicador de tasa'] = indicadorDeTasa
archivoPlano['monedaDocumento'] = monedaDelDocumento
archivoPlano['monedaFinal'] = monedaBaseDeConversion
archivoPlano['Moneda local'] = monedaLocal
archivoPlano['solicitante'] = solicitante
archivoPlano['Tasa de conversión'] = tasaDeConversion
archivoPlano['Tasa local'] = tasaLocal
archivoPlano['Indicador de contacto'] = indicadorDeContacto
archivoPlano['motivo'] = motivo
archivoPlano['Indicador de obsequio'] = indicadorDeObsequio
archivoPlano['notas'] = notas
archivoPlano['Detalle'] = detalle
archivoPlano['Descripción del item'] = descripcionDelItem
archivoPlano['unidadNegocio'] = unidadNegocioMovimiento
archivoPlano['Tasa de descuento condicionado'] = tasaDescuentoCondicionado
archivoPlano['fechaDocumento'] = fechaDescarga

#Cambiar después eliminando columnas innecesarias
archivoPlano['fechaEntrega'] = fechaEntrega
archivoPlano['fechaEntregaLejanias'] = fechaEntregaLejanias
archivoPlano['Descuento global 1'] = descuentoGlobal1
archivoPlano['Descuento global 2'] = descuentoGlobal2


proveedores = get_excel_sh(siteAprovisionamiento,'Indicadores','Agroquímicos',"Parámetro",'Diccionarios.xlsx','Proveedores',1)
proveedores = proveedores[proveedores["Habilitado"] == "Si"]
proveedores = proveedores.sort_values(by = ['Razón social'], ascending = [True],ignore_index=True )
proveedores = proveedores.groupby(['Código','Razón social','Condicion de pago'],as_index=False).agg({'Sucursal':'min'})
proveedores['Sucursal'] = proveedores['Sucursal'].astype(str).str.zfill(3) #Verificar después con esas sucursales raras o con más de dos dígitos
archivoPlano = pd.merge(archivoPlano,proveedores[['Razón social','Código','Condicion de pago','Sucursal']],how='left',left_on= ['Razón social proveedor'], right_on= ['Razón social'])

folder = siteAprovisionamiento.Folder(f'Shared%20Documents/Indicadores/Agroquímicos')
bodegas = pd.read_excel(folder.get_file('Diccionarios.xlsx'), sheet_name='Bodegas', dtype = {'centroOperacion':str})
archivoPlano = pd.merge(archivoPlano,bodegas[['centroOperacion','Bodega']],how='left',left_on= ['Bodega'], right_on= ['Bodega'])
archivoPlano.rename(columns = {'Código':'proveedor','Sucursal':'sucursal','Condicion de pago':'condicionPago','Bodega':'bodega','UM Compras':'unidadMedida','Unidades de compra':'cantidadPedida','Precio Actual Compra':'precioUnitario','SisFinCode':'item'}, inplace = True)


listbodegas = ['IN001', 'IN008', 'IN013', 'IN052', 'IN054', 'IN055', 'IN062', 'IN063', 'IN071', 'IN078', 'IN083', 'IN155']
archivoPlano.loc[archivoPlano['bodega'].isin(listbodegas), 'fechaEntrega'] = fechaEntregaLejanias
archivoPlano['centroOperacion'] = archivoPlano['centroOperacion'].astype(str).str.zfill(3) #Verificar después con esas sucursales raras o con más de dos dígitos

documentos = archivoPlano.drop_duplicates(subset=['consecutivoOC'])
documentos = documentos[['centroOperacion','tipoDocumento','consecutivoOC','fechaDocumento','comprador','proveedor','sucursal','condicionPago','monedaDocumento','monedaFinal','notas','documentoReferencia','solicitante']]
archivoPlano['centroOperacionM'] = archivoPlano['centroOperacion']

#------Ajustar fecha de los bichos
semanasDeAbastecimiento = get_excel_sh(siteAprovisionamiento,'Indicadores','Agroquímicos',"Parámetro",'Diccionarios.xlsx','Fincas',1)
archivoPlano = pd.merge(archivoPlano,semanasDeAbastecimiento[['Bodega','Semanas de abastecimiento']],how='left',left_on= ['bodega'], right_on= ['Bodega'])

#Colocarle la semana que corresponde a los productos de bichitos a las bodegas que se abastecen a una semana
indice = 0
while indice<len(archivoPlano):
    item = archivoPlano['item'][indice]
    semanaAbastecimiento = archivoPlano['Semanas de abastecimiento'][indice]
    if item in listBichitos:
        if semanaAbastecimiento == 1:
            archivoPlano['Semana'][indice] = semana+2
    indice+=1

calendario = get_excel_sh(siteDBLogistics,"Calendarios",'','','calendarioSunshine.xlsx','Hoja1',4)
calendario = calendario[calendario["año"] == año]
calendario = calendario[calendario['Dia semana'].isin(['Friday', 'viernes'])]
palabraSemana = 'Semana'
calendario['SemanaTexto'] = palabraSemana + calendario['semana'].astype(str)

archivoPlano = pd.merge(archivoPlano,calendario[['Fecha','semana','año','SemanaTexto']],how='left',left_on= ['Semana'], right_on= ['semana'])
archivoPlano['Fecha'] = archivoPlano['Fecha'].fillna(today)
archivoPlano[['Fecha']] = archivoPlano[["Fecha"]].astype(str)
archivoPlano['Fecha'] = archivoPlano['Fecha'].apply(lambda x:dateToString(str(x)))

for i in listBichitos:
    archivoPlano['fechaEntrega'] = np.where(archivoPlano['item'] == i,archivoPlano['Fecha'],archivoPlano['fechaEntrega'])
    #archivoPlano['fechaEntrega'] = np.where(archivoPlano['item'] == i,archivoPlano['Fecha'],archivoPlano['fechaEntregaLejanias'])
    archivoPlano['notas'] = np.where(archivoPlano['item'] == i,archivoPlano['SemanaTexto'],archivoPlano['notas'])
    
#Columnas finales
movimientos = archivoPlano[['centroOperacion','tipoDocumento','consecutivoOC','consecutivoItem','bodega','motivo','centroOperacionM','unidadMedida','cantidadPedida','fechaEntrega','precioUnitario','notas','item','unidadNegocio']]

#-----------------------Archivos planos en Excel
archivoPlanoExcelDoc = documentos.copy()
archivoPlanoExcelDoc = pd.merge(archivoPlanoExcelDoc,proveedores[['Código','Razón social']],how='left',left_on= ['proveedor'], right_on= ['Código'])
archivoPlanoExcelDoc = archivoPlanoExcelDoc[['consecutivoOC','fechaDocumento','comprador','Razón social','sucursal','condicionPago','monedaDocumento','solicitante']]
archivoPlanoExcelDoc.rename(columns = {'consecutivoOC':'ConsecutivoOC','fechaDocumento':'Fecha documento','comprador':'Comprador','sucursal':'Sucursal','condicionPago':'Condición pago','monedaDocumento':'Moneda documento','solicitante':'Solicitante'}, inplace = True)

bodegasIN = get_excel_sh(siteAprovisionamiento,'Indicadores','Agroquímicos',"Parámetro",'Diccionarios.xlsx','Fincas',1)
archivoPlanoExcelMov = movimientos.copy()
archivoPlanoExcelMov = pd.merge(archivoPlanoExcelMov,bodegasIN[['Bodega','Descripción Bodega']],how='left',left_on= ['bodega'], right_on= ['Bodega'])
archivoPlanoExcelMov = pd.merge(archivoPlanoExcelMov,archivoPlanoCopia[['SisFinCode','Quimico']],how='left',left_on= ['item'], right_on= ['SisFinCode'])
archivoPlanoExcelMov = archivoPlanoExcelMov[['Descripción Bodega','tipoDocumento','consecutivoOC','consecutivoItem','unidadMedida','cantidadPedida','fechaEntrega','precioUnitario','Quimico']]
archivoPlanoExcelMov.rename(columns = {'Descripción Bodega':'Descripción bodega','tipoDocumento':'Tipo documento','consecutivoOC':'ConsecutivoOC','consecutivoItem':'Consecutivo Item','unidadMedida':'Unidad de medida','cantidadPedida':'Cantidad pedida','fechaEntrega':'Fecha entrega','precioUnitario':'Precio','Quimico':'Descripción item'}, inplace = True)

if adicionales==1:
    create_excel(documentos,f"ArchivoPlanoSemana{semana}","Documentos")
    create_sheet(movimientos,f"ArchivoPlanoSemana{semana}",'Movimientos')

    create_excel(archivoPlanoExcelDoc,f"OrdenCompraSemana{semana}","Documentos")
    create_sheet(archivoPlanoExcelMov,f"OrdenCompraSemana{semana}",'Movimientos')
else:
    create_excel(documentos,f"ArchivoPlanoSemana{semana}Adicionales","Documentos")
    create_sheet(movimientos,f"ArchivoPlanoSemana{semana}Adicionales",'Movimientos')

    create_excel(archivoPlanoExcelDoc,f"OrdenCompraSemana{semana}Adicionales","Documentos")
    create_sheet(archivoPlanoExcelMov,f"OrdenCompraSemana{semana}Adicionales",'Movimientos')    

exit()
if adicionales==1:
    file_upload_to_sharepoint(siteAprovisionamiento,año,f'Semana{semana}',f"ArchivoPlanoSemana{semana}")
    file_upload_to_sharepoint(siteAprovisionamiento,año,f'Semana{semana}',f"OrdenCompraSemana{semana}")
else:
    file_upload_to_sharepoint(siteAprovisionamiento,año,f'Semana{semana}/Adicionales',f"ArchivoPlanoSemana{semana}Adicionales")
    file_upload_to_sharepoint(siteAprovisionamiento,año,f'Semana{semana}/Adicionales',f"OrdenCompraSemana{semana}Adicionales")

#Time
producto = f'Fase 4: Generación de archivo plano'
elapsed_time = time() - start_time
print("Tiempo: %.10f segundos." % elapsed_time)
input(f'{producto} finalizado!... Presione enter para salir. ')